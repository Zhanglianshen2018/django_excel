#!/usr/bin/env python
# coding: utf-8

# In[53]:



import os
import sys
import csv
#import easygui as g
from openpyxl import load_workbook


'''
def ui():
    if g.ccbox(msg='请选择文件夹：', title='数据读取 ', choices=('打开', '取消')):
            file_dir=g.diropenbox(msg='请选择文件夹',title='文件夹选择')
            if file_dir:
                return file_dir
            else:
                sys.exit(0)
    else:
        sys.exit(0)
'''

def operate_sheet(ws):
    '''
    car={}
    car['车型']=ws['D3'].value
    car['VIN']=ws['M3'].value
    car['测试者']=ws['U3'].value[4:7].strip()
    car['测试日期']=ws['U3'].value[-8:]
    
    car['后外倾角左']=(int(ws['V5'].value)+int(ws['X5'].value)/60)*(-1 if ws['U5'].value=='－'else 1)
    car['后外倾角右']=(int(ws['V6'].value)+int(ws['X6'].value)/60)*(-1 if ws['U6'].value=='－'else 1)
    car['前外倾角左']=(int(ws['V11'].value)+int(ws['X11'].value)/60)*(-1 if ws['U11'].value=='－'else 1)
    car['前外倾角右']=(int(ws['V12'].value)+int(ws['X12'].value)/60)*(-1 if ws['U12'].value=='－'else 1)
    '''
    
    car=[]
    car.append(ws['D3'].value)
    car.append(ws['M3'].value)
    car.append(ws['U3'].value[4:7].strip())
    car.append(ws['U3'].value[-8:])
    
    car.append((int(ws['V5'].value)+int(ws['X5'].value)/60)*(-1 if ws['U5'].value=='－'else 1))
    car.append((int(ws['V6'].value)+int(ws['X6'].value)/60)*(-1 if ws['U6'].value=='－'else 1))
    car.append((int(ws['V11'].value)+int(ws['X11'].value)/60)*(-1 if ws['U11'].value=='－'else 1))
    car.append((int(ws['V12'].value)+int(ws['X12'].value)/60)*(-1 if ws['U12'].value=='－'else 1))

    return car

            
def operate_excel(temp_path):
    wb = load_workbook(temp_path)
    result_list=[]
    
    for sheet in wb:
        result=operate_sheet(sheet)
        result_list.append(result)
    return result_list

def write_to_csv(columns,results):
    with open('results.csv','w',newline='') as f:
        writer = csv.writer(f)

        #先写入columns_name
        writer.writerow(columns)
        #写入多行用writerows
        writer.writerows(results)


# In[54]:

if __name__ == '__main__':
    file_dir=ui()
    excel_list=os.listdir(file_dir)
    results=[]

    for file in excel_list:
        temp_path=os.path.join(file_dir,file)
        results=results+operate_excel(temp_path)

    columns=['车型','VIN','测试者','测试日期','后外倾角左','后外倾角右','前外倾角左','前外倾角右']
    write_to_csv(columns,results)
    #g.msgbox("完成")




